"""
report_xlsx.py
---------------
Builds the reconciliation report as an .xlsx workbook instead of a PDF.
This is the readable/filterable alternative to report.py: every sheet is
a normal table you can sort, filter, and search in Excel, rather than a
fixed page layout.

Sheets produced:
  Summary      - counts by status
  By Field     - counts by field x status
  Salary Analysis - month-by-month gross salary, LOP days/deduction, net salary
  Issues       - genuine mismatches, missing values, and missing keys only
  Accepted Transforms - values accepted after the configured normalization
  Full Detail  - every (record, field) checked, for audit trail
"""

from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STATUS_FILL = {
    "MATCH": "D9EAD3",
    "MATCH_AFTER_TRANSFORM": "FFF2CC",
    "MISMATCH": "F4CCCC",
    "MISSING_IN_DESTINATION": "F9CB9C",
    "FIELD_NOT_IN_SOURCE": "D9D2E9",
    "KEY_NOT_FOUND_IN_SOURCE": "EA9999",
}

STATUS_LABEL = {
    "MATCH": "Match",
    "MATCH_AFTER_TRANSFORM": "Match (after expected transform)",
    "MISMATCH": "Mismatch",
    "MISSING_IN_DESTINATION": "Missing in destination",
    "FIELD_NOT_IN_SOURCE": "Field not present in any source",
    "KEY_NOT_FOUND_IN_SOURCE": "Key not found in any source",
}

SEVERITY_ORDER = [
    "KEY_NOT_FOUND_IN_SOURCE",
    "MISMATCH",
    "MISSING_IN_DESTINATION",
    "FIELD_NOT_IN_SOURCE",
]

HEADER_FILL = PatternFill("solid", fgColor="434343")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial")
BODY_FONT = Font(name="Arial", size=10)


def _fmt_source_values(source_values: dict) -> str:
    if not source_values:
        return "-"
    parts = []
    for sname, vals in source_values.items():
        uniq = sorted(set(str(v) for v in vals))
        parts.append(f"{sname}: {', '.join(uniq)}")
    return "\n".join(parts)


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_detail_sheet(wb, sheet_name, rows, statuses):
    ws = wb.create_sheet(sheet_name)
    header = ["Key", "Field", "Status", "Destination Value", "Source Value(s)"]
    ws.append(header)
    for row, status in zip(rows, statuses):
        ws.append(row)
        r = ws.max_row
        if status:
            fill = PatternFill("solid", fgColor=STATUS_FILL[status])
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill
                ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    _style_header(ws, len(header))
    _autosize(ws, [10, 16, 30, 22, 45])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"
    return ws


def _write_salary_analysis_sheet(wb, salary_analysis):
    ws = wb.create_sheet("Salary Analysis")
    header = [
        "Employee ID", "Month", "Days in Month", "Eligible Days",
        "Loss of Pay Days", "Gross Salary", "Loss of Pay Deduction", "Net Salary",
    ]
    ws.append(header)
    for item in salary_analysis:
        ws.append([
            item.get("EmployeeID"), item.get("Month"), item.get("DaysInMonth"),
            item.get("EligibleDays"), item.get("LossOfPayDays"),
            item.get("GrossSalary"), item.get("LossOfPayDeduction"),
            item.get("NetSalary"),
        ])
        for c in range(1, len(header) + 1):
            ws.cell(row=ws.max_row, column=c).font = BODY_FONT
    _style_header(ws, len(header))
    _autosize(ws, [14, 10, 14, 14, 18, 18, 24, 18])
    if ws.max_row > 1:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=5).number_format = "0.00"
            for col in (6, 7, 8):
                ws.cell(row=row, column=col).number_format = "#,##0.00"
        ws.auto_filter.ref = f"A1:H{ws.max_row}"
    return ws


def build_report(results: list, summary: dict, config: dict, output_path: str,
                  title: str = "Source-to-Destination Reconciliation Report",
                  salary_analysis: list | None = None):
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- Summary sheet ----------
    ws = wb.create_sheet("Summary")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A2"] = (
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Key field: {config['key_field']}  |  "
        f"Sources: {', '.join(s['name'] for s in config['sources'])}"
    )
    ws["A2"].font = Font(italic=True, size=9, name="Arial")

    ws.append([])
    ws.append(["Status", "Count"])
    header_row = ws.max_row
    status_order = ["MATCH", "MATCH_AFTER_TRANSFORM", "MISMATCH",
                     "MISSING_IN_DESTINATION", "FIELD_NOT_IN_SOURCE",
                     "KEY_NOT_FOUND_IN_SOURCE"]
    for status in status_order:
        if status in summary:
            ws.append([STATUS_LABEL[status], summary[status]])
            fill = PatternFill("solid", fgColor=STATUS_FILL[status])
            for c in (1, 2):
                ws.cell(row=ws.max_row, column=c).fill = fill
                ws.cell(row=ws.max_row, column=c).font = BODY_FONT
    ws.append(["TOTAL FIELD CHECKS", summary.get("TOTAL_CHECKS", 0)])
    for c in (1, 2):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True, name="Arial")
    _style_header(ws, 2, row=header_row)
    _autosize(ws, [32, 12])

    # ---------- By Field sheet ----------
    ws2 = wb.create_sheet("By Field")
    by_field = defaultdict(lambda: defaultdict(int))
    for r in results:
        by_field[r["field"]][r["status"]] += 1

    all_fields = list(config["fields_to_validate"]) + [
        spec["output_field"] for spec in config.get("computed_fields", [])
    ]
    field_order = {field: position for position, field in enumerate(all_fields)}
    ws2.append(["Field", "Match", "Match (transform)", "Mismatch",
                "Missing in Dest", "Not in Source", "Key Not Found"])
    for field in all_fields:
        counts = by_field[field]
        ws2.append([
            field,
            counts.get("MATCH", 0),
            counts.get("MATCH_AFTER_TRANSFORM", 0),
            counts.get("MISMATCH", 0),
            counts.get("MISSING_IN_DESTINATION", 0),
            counts.get("FIELD_NOT_IN_SOURCE", 0),
            counts.get("KEY_NOT_FOUND_IN_SOURCE", 0),
        ])
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 8):
            ws2.cell(row=r, column=c).font = BODY_FONT
    _style_header(ws2, 7)
    _autosize(ws2, [18, 9, 16, 10, 15, 14, 13])
    ws2.auto_filter.ref = f"A1:G{ws2.max_row}"

    # ---------- Salary Analysis sheet ----------
    if salary_analysis:
        _write_salary_analysis_sheet(wb, salary_analysis)

    # ---------- Issues sheet ----------
    issues = [r for r in results if r["status"] in SEVERITY_ORDER]
    issues.sort(key=lambda r: (
        SEVERITY_ORDER.index(r["status"]), r["key"] or "",
        field_order.get(r["field"], len(field_order)),
    ))
    rows, statuses = [], []
    for r in issues:
        rows.append([
            r["key"], r["field"], STATUS_LABEL[r["status"]],
            r["destination_value"] if r["destination_value"] not in (None, "") else "(empty)",
            _fmt_source_values(r["source_values"]),
        ])
        statuses.append(r["status"])
    _write_detail_sheet(wb, "Issues", rows, statuses)

    # ---------- Accepted transformations sheet ----------
    # These are successful validations, not errors. Keeping them separate
    # makes filler values such as XX and expected formatting normalizations
    # auditable without polluting the Issues sheet.
    transformed = [r for r in results if r["status"] == "MATCH_AFTER_TRANSFORM"]
    transformed.sort(key=lambda r: (
        r["key"] or "", field_order.get(r["field"], len(field_order))
    ))
    rows, statuses = [], []
    for r in transformed:
        rows.append([
            r["key"], r["field"], STATUS_LABEL[r["status"]],
            r["destination_value"] if r["destination_value"] not in (None, "") else "(empty)",
            _fmt_source_values(r["source_values"]),
        ])
        statuses.append(r["status"])
    _write_detail_sheet(wb, "Accepted Transforms", rows, statuses)

    # ---------- Full Detail sheet ----------
    all_sorted = sorted(results, key=lambda r: (
        r["key"] or "", field_order.get(r["field"], len(field_order))
    ))
    rows, statuses = [], []
    for r in all_sorted:
        rows.append([
            r["key"], r["field"], STATUS_LABEL[r["status"]],
            r["destination_value"] if r["destination_value"] not in (None, "") else "(empty)",
            _fmt_source_values(r["source_values"]),
        ])
        statuses.append(r["status"])
    _write_detail_sheet(wb, "Full Detail", rows, statuses)

    wb.move_sheet("Summary", offset=-len(wb.sheetnames))
    wb.active = 0
    wb.save(output_path)
