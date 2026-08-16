"""
readers.py
----------
Turns heterogeneous source files (CSV, XML, flat/notepad text, Access) and
destination files (CSV, XML, or a database) into one common shape:

    [ {canonical_field_1: value, canonical_field_2: value, ...}, ... ]

Every reader takes a source's config block (from config.json) and returns
records with keys already renamed to the CANONICAL field names, using each
source's own field_map.

Some sources carry MULTIPLE raw rows per key. There are two different ways
to collapse those down to one row per key, depending on what the extra
rows mean:

  "aggregate" -- the rows are components that should be SUMMED (e.g. a
  bonus file with one row per quarter per employee):
    "aggregate": {
        "group_by": "EmployeeID",
        "sum_fields": {"BonusAmount": "TotalBonus"}
    }

  "dedupe" -- the rows are the SAME field re-stated over time (e.g. an
  employee transferred departments or was rehired, so the source has one
  row per event) and only the most recent one should count:
    "dedupe": {
        "group_by": "EmployeeID",
        "keep_latest_by": "EffectiveDate"
    }

A source uses at most one of these. Rows that fail to parse (wrong column
count, unparseable date, etc.) are skipped with a warning collected in
READ_WARNINGS rather than crashing the whole run -- real extracts are
rarely perfectly clean.
"""

import csv
import json
import re
import xml.etree.ElementTree as ET
from datetime import datetime

from transforms import numeric

# Populated during read_source() calls; main.py prints these after loading
# all sources so a few bad rows don't silently disappear.
READ_WARNINGS = []


def _parse_date(v):
    """Best-effort date parsing across the handful of formats real extracts
    tend to use. Returns None (never raises) if nothing matches."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _apply_field_map(raw_record: dict, field_map: dict) -> dict:
    """Rename raw source column names -> canonical field names.
    Columns not listed in field_map are dropped (not relevant to reconciliation)."""
    out = {}
    for raw_key, canonical_key in field_map.items():
        # raw_key lookup is case/whitespace tolerant
        val = None
        for k, v in raw_record.items():
            if k is not None and k.strip().lower() == raw_key.strip().lower():
                val = v
                break
        out[canonical_key] = val.strip() if isinstance(val, str) else val
    return out


def read_csv_source(cfg: dict) -> list:
    path = cfg["path"]
    delimiter = cfg.get("delimiter", ",")
    records = []
    with open(path, newline="", encoding=cfg.get("encoding", "utf-8-sig")) as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for line_num, row in enumerate(reader, start=2):
            if None in row or any(v is None for v in row.values()):
                READ_WARNINGS.append(
                    f"{cfg.get('name', path)}: line {line_num} has the wrong number of "
                    f"columns -- skipped ({row!r})"
                )
                continue
            records.append(row)
    return [_apply_field_map(r, cfg["field_map"]) for r in records]


def read_flat_source(cfg: dict) -> list:
    """Handles 'Notepad'-style flat files. Two modes:
      mode='delimited'  -> like CSV but arbitrary delimiter, optional header
      mode='keyvalue'   -> records separated by blank lines, each line 'Key: Value'
    """
    mode = cfg.get("mode", "delimited")
    path = cfg["path"]

    if mode == "keyvalue":
        records = []
        current = {}
        with open(path, encoding=cfg.get("encoding", "utf-8")) as f:
            for line in f:
                line = line.strip()
                if not line:
                    if current:
                        records.append(current)
                        current = {}
                    continue
                if ":" in line:
                    k, v = line.split(":", 1)
                    current[k.strip()] = v.strip()
            if current:
                records.append(current)
        return [_apply_field_map(r, cfg["field_map"]) for r in records]

    # delimited mode
    delimiter = cfg.get("delimiter", "|")
    has_header = cfg.get("has_header", True)
    with open(path, encoding=cfg.get("encoding", "utf-8")) as f:
        lines = [l.rstrip("\n") for l in f if l.strip() != ""]
    if has_header:
        header = [h.strip() for h in lines[0].split(delimiter)]
        rows = lines[1:]
    else:
        header = cfg["columns"]  # must be supplied in config if no header row
        rows = lines
    records = []
    for line_num, row in enumerate(rows, start=(2 if has_header else 1)):
        parts = [p.strip() for p in row.split(delimiter)]
        if len(parts) != len(header):
            READ_WARNINGS.append(
                f"{cfg.get('name', path)}: line {line_num} has {len(parts)} field(s), "
                f"expected {len(header)} -- skipped ({row!r})"
            )
            continue
        records.append(dict(zip(header, parts)))
    return [_apply_field_map(r, cfg["field_map"]) for r in records]


def read_xml_source(cfg: dict) -> list:
    path = cfg["path"]
    record_xpath = cfg["record_xpath"]  # e.g. ".//Employee"
    tree = ET.parse(path)
    root = tree.getroot()
    records = []
    for node in root.findall(record_xpath):
        raw = {}
        # attributes
        raw.update(node.attrib)
        # child text elements
        for child in node:
            raw[child.tag] = child.text.strip() if child.text else child.text
        records.append(raw)
    return [_apply_field_map(r, cfg["field_map"]) for r in records]


def read_access_source(cfg: dict) -> list:
    """Reads an MS Access .mdb/.accdb table.
    Requires either `pyodbc` with the MS Access driver (Windows), or
    `mdbtools` + `pandas_access` on Linux. Kept isolated so the rest of the
    pipeline works even if this dependency isn't installed."""
    path = cfg["path"]
    table = cfg["table"]
    try:
        import pandas_access as mdb  # linux, needs mdbtools installed
        df = mdb.read_table(path, table)
    except ImportError:
        try:
            import pyodbc
            import pandas as pd
            conn_str = (
                r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
                rf"DBQ={path};"
            )
            conn = pyodbc.connect(conn_str)
            df = pd.read_sql(f"SELECT * FROM [{table}]", conn)
        except Exception as e:
            raise RuntimeError(
                "Could not read Access file. Install `mdbtools` + `pandas_access` "
                "(Linux/Mac) or use pyodbc with the Access driver (Windows). "
                f"Original error: {e}"
            )
    records = df.to_dict(orient="records")
    return [_apply_field_map(r, cfg["field_map"]) for r in records]


def read_excel_source(cfg: dict) -> list:
    """Reads a .xlsx/.xlsm sheet. One row per record (e.g. one row per
    employee with a column per month of salary)."""
    from openpyxl import load_workbook

    path = cfg["path"]
    sheet_name = cfg.get("sheet")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    records = []
    for row in rows:
        if row is None or all(v is None for v in row):
            continue
        raw = dict(zip(header, row))
        records.append(raw)
    return [_apply_field_map(r, cfg["field_map"]) for r in records]


READERS = {
    "csv": read_csv_source,
    "flat": read_flat_source,
    "notepad": read_flat_source,
    "xml": read_xml_source,
    "access": read_access_source,
    "excel": read_excel_source,
    "xlsx": read_excel_source,
}


def apply_aggregation(records: list, agg_cfg: dict) -> list:
    """Collapses multiple raw rows sharing the same group_by key into one
    row, summing each field listed in sum_fields. Runs after field_map, so
    group_by/sum_fields use CANONICAL field names."""
    if not agg_cfg:
        return records
    group_by = agg_cfg["group_by"]
    sum_fields = agg_cfg.get("sum_fields", {})

    groups = {}
    order = []
    for rec in records:
        key = rec.get(group_by)
        if key not in groups:
            groups[key] = {"_first": rec, "_totals": {f: 0.0 for f in sum_fields}, "_seen": {f: False for f in sum_fields}}
            order.append(key)
        g = groups[key]
        for raw_field, out_field in sum_fields.items():
            n = numeric(rec.get(raw_field))
            if n is not None:
                g["_totals"][raw_field] += n
                g["_seen"][raw_field] = True

    collapsed = []
    for key in order:
        g = groups[key]
        rec = dict(g["_first"])  # carries over any non-summed fields too
        for raw_field, out_field in sum_fields.items():
            rec[out_field] = round(g["_totals"][raw_field], 2) if g["_seen"][raw_field] else None
        collapsed.append(rec)
    return collapsed


def apply_dedupe(records: list, dedupe_cfg: dict) -> list:
    """Collapses multiple raw rows sharing the same group_by key down to
    ONE row: the one with the latest keep_latest_by date. Use this (instead
    of aggregate) when extra rows represent the same person restated over
    time -- a transfer, a rehire, a correction -- rather than components to
    sum. Rows whose date fails to parse are treated as oldest (kept only if
    nothing else exists for that key), and a warning is recorded."""
    if not dedupe_cfg:
        return records
    group_by = dedupe_cfg["group_by"]
    date_field = dedupe_cfg["keep_latest_by"]

    best = {}
    order = []
    for rec in records:
        key = rec.get(group_by)
        parsed = _parse_date(rec.get(date_field))
        if parsed is None:
            READ_WARNINGS.append(
                f"dedupe on '{group_by}'={key!r}: could not parse '{date_field}'="
                f"{rec.get(date_field)!r}, treating as earliest"
            )
        if key not in best:
            best[key] = (parsed, rec)
            order.append(key)
        else:
            existing_date, _ = best[key]
            if parsed is not None and (existing_date is None or parsed > existing_date):
                best[key] = (parsed, rec)
    return [best[k][1] for k in order]


def read_source(cfg: dict) -> list:
    reader = READERS.get(cfg["type"])
    if not reader:
        raise ValueError(f"Unsupported source type: {cfg['type']}")
    records = reader(cfg)
    records = apply_aggregation(records, cfg.get("aggregate"))
    records = apply_dedupe(records, cfg.get("dedupe"))
    return records


def read_destination(cfg: dict) -> list:
    """Destination can be csv, xml, flat, or a database query result already
    exported to one of those forms (simplest + most portable), or a live DB
    via SQLAlchemy connection string."""
    if cfg["type"] in READERS:
        return READERS[cfg["type"]](cfg)
    if cfg["type"] == "database":
        import sqlalchemy as sa
        engine = sa.create_engine(cfg["connection_string"])
        with engine.connect() as conn:
            result = conn.execute(sa.text(cfg["query"]))
            rows = [dict(r._mapping) for r in result]
        return [_apply_field_map(r, cfg["field_map"]) for r in rows]
    raise ValueError(f"Unsupported destination type: {cfg['type']}")
